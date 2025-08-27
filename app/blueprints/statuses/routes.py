import io
from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required
from openpyxl import load_workbook
from ...extensions import db
from ...models import EquipmentStatus
from ...forms.statuses import EquipmentStatusForm

statuses_bp = Blueprint("statuses", __name__)

@statuses_bp.route("/", methods=["GET"])
@login_required
def list_():
    rows = EquipmentStatus.query.order_by(EquipmentStatus.name.asc()).all()
    return render_template("statuses/list.html", rows=rows)

@statuses_bp.route("/novo", methods=["GET", "POST"])
@login_required
def create():
    form = EquipmentStatusForm()
    if form.validate_on_submit():
        s = EquipmentStatus(name=form.name.data.strip(), color=(form.color.data or '').strip() or None)
        db.session.add(s)
        db.session.commit()
        flash("Status salvo.", "success")
        return redirect(url_for("statuses.list_"))
    return render_template("statuses/form.html", form=form, mode="create")

@statuses_bp.route("/<int:status_id>/editar", methods=["GET", "POST"])
@login_required
def edit(status_id):
    s = EquipmentStatus.query.get_or_404(status_id)
    form = EquipmentStatusForm(obj=s)
    if form.validate_on_submit():
        form.populate_obj(s)
        db.session.commit()
        flash("Status atualizado.", "success")
        return redirect(url_for("statuses.list_"))
    return render_template("statuses/form.html", form=form, mode="edit", row=s)

@statuses_bp.route("/importar", methods=["GET", "POST"])
@login_required
def import_():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('Selecione um arquivo .xlsx', 'error')
            return redirect(request.url)
        try:
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            if 'Sum' not in wb.sheetnames:
                flash("A aba 'Sum' não foi encontrada.", 'error')
                return redirect(request.url)
            ws = wb['Sum']
            header = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if 'status' not in header:
                flash("A aba 'Sum' deve conter a coluna 'Status' na primeira linha.", 'error')
                return redirect(request.url)
            col = header.index('status') + 1
            seen = set()
            created = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col-1]
                if not val: continue
                name = str(val).strip()
                key = name.lower()
                if key in seen: continue
                seen.add(key)
                if not EquipmentStatus.query.filter(EquipmentStatus.name.ilike(name)).first():
                    db.session.add(EquipmentStatus(name=name))
                    created += 1
            db.session.commit()
            flash(f"Importação concluída. {created} status adicionados.", 'success')
            return redirect(url_for('statuses.list_'))
        except Exception as e:
            flash(f"Falha ao importar: {e}", 'error')
            return redirect(request.url)
    return render_template('statuses/import.html')
