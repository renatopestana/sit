import io, re, unicodedata
from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required
from openpyxl import load_workbook
from ...extensions import db
from ...models import Equipment, User, Client, Project, EquipmentStatus
from ...forms.equipment import EquipmentForm

inventory_bp = Blueprint("inventory", __name__)

@inventory_bp.route("/", methods=["GET"])
@login_required
def list_():
    q = request.args.get("q", "").strip()
    query = Equipment.query
    if q:
        like = f"%{q}%"
        query = query.filter((Equipment.name.ilike(like)) | (Equipment.pn.ilike(like)) | (Equipment.serial_number.ilike(like)))
    rows = query.order_by(Equipment.created_at.desc()).all()
    return render_template("inventory/list.html", rows=rows, q=q)

@inventory_bp.route("/novo", methods=["GET", "POST"])
@login_required
def create():
    form = EquipmentForm()
    _fill_choices(form)
    if form.validate_on_submit():
        e = Equipment(
            name=form.name.data,
            pn=form.pn.data or None,
            model_number=form.model_number.data or None,
            serial_number=form.serial_number.data or None,
            machine_installed=form.machine_installed.data or None,
            image_ref=form.image_ref.data or None,
            asset_tag=form.asset_tag.data or form.pn.data or None,
            category=form.category.data or None,
            brand=form.brand.data or None,
            owner_id=form.owner_id.data or None,
            current_responsible_id=form.current_responsible_id.data or None,
            location_id=form.location_id.data or None,
            project_id=form.project_id.data or None,
            status_id=form.status_id.data or None,
            notes=form.notes.data or None,
        )
        db.session.add(e)
        db.session.commit()
        flash("Equipamento salvo.", "success")
        return redirect(url_for("inventory.list_"))
    return render_template("inventory/form.html", form=form, mode="create")

@inventory_bp.route("/<int:equipment_id>/editar", methods=["GET", "POST"])
@login_required
def edit(equipment_id):
    e = Equipment.query.get_or_404(equipment_id)
    form = EquipmentForm(obj=e)
    _fill_choices(form)
    if form.validate_on_submit():
        form.populate_obj(e)
        db.session.commit()
        flash("Equipamento atualizado.", "success")
        return redirect(url_for("inventory.list_"))
    return render_template("inventory/form.html", form=form, mode="edit", row=e)

# Helpers para importação
_SPLIT_RE = re.compile(r"\s*(?:/|&|,|;| e )\s*", re.IGNORECASE)

def _normalize_name(s: str) -> str:
    s = (s or '').strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)

def _slug_email(name: str) -> str:
    n = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    n = re.sub(r"[^a-zA-Z0-9\s]", "", n).strip().lower()
    n = re.sub(r"\s+", ".", n)
    base = n or 'user'
    email = f"{base}@autogen.local"
    i = 1
    while User.query.filter_by(email=email).first():
        email = f"{base}{i}@autogen.local"; i += 1
    return email

def _get_or_create_user_by_fullname(full_name: str, counters: dict):
    if not full_name:
        return None
    u = User.query.filter(db.func.lower(User.full_name)==full_name.lower()).first()
    if u:
        return u
    u = User(full_name=_normalize_name(full_name), email=_slug_email(full_name))
    u.set_password('ChangeMe123!')
    db.session.add(u)
    db.session.flush()
    counters['users'] = counters.get('users', 0) + 1
    return u

def _get_or_create_client_by_name(nome_razao: str, counters: dict):
    if not nome_razao:
        return None
    c = Client.query.filter(db.func.lower(Client.nome_razao)==nome_razao.lower()).first()
    if c:
        return c
    c = Client(tipo='PJ', nome_razao=_normalize_name(nome_razao), endereco='Criado automaticamente pelo importador')
    db.session.add(c)
    db.session.flush()
    counters['clients'] = counters.get('clients', 0) + 1
    return c

@inventory_bp.route('/importar', methods=['GET','POST'])
@login_required
def import_():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('Selecione um arquivo .xlsx', 'error')
            return redirect(request.url)
        try:
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            # 1) Importar Status da aba "Sum" (se existir)
            if 'Sum' in wb.sheetnames:
                ws = wb['Sum']
                header = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                if 'status' in header:
                    col = header.index('status') + 1
                    seen = set()
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        val = row[col-1]
                        if not val: continue
                        name = str(val).strip()
                        key = name.lower()
                        if key in seen: continue
                        seen.add(key)
                        if not EquipmentStatus.query.filter(db.func.lower(EquipmentStatus.name)==name.lower()).first():
                            db.session.add(EquipmentStatus(name=name))
                    db.session.commit()
            # 2) Importar equipamentos da planilha principal (primeira aba)
            ws = wb.active
            # detectar cabeçalho
            header_row = None
            for r in range(1, 6):
                vals = [str(c.value).strip() if c.value is not None else '' for c in ws[r]]
                if any(v.lower()=='item' for v in vals):
                    header_row = r
                    header_vals = vals
                    break
            if not header_row:
                flash('Cabeçalho não localizado (coluna Item).', 'error')
                return redirect(request.url)
            index = {v.lower(): i for i, v in enumerate(header_vals)}
            def col(*aliases):
                for a in aliases:
                    if a.lower() in index: return index[a.lower()]
                return None
            c_item = col('Item')
            c_pn = col('PN')
            c_model = col('Model Number','Model')
            c_sn = col('SN','Serial Number')
            c_loc = col('Location')
            c_mach = col('Machine Installed')
            c_status = col('Status')
            c_proj = col('Project')
            c_owner = col('Owner')
            c_resp = col('Current Responsible','Current Reponsible')
            c_obs = col('Obs','Observacao','Observações')
            c_img = col('Imagem de Referência','Image','Image Ref')

            created, updated = 0, 0
            missing_users = set()
            counters = {}

            for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                def get(c):
                    return str(row[c]).strip() if (c is not None and row[c] is not None) else ''

                name = get(c_item)
                if not name: continue
                pn = get(c_pn)
                model_number = get(c_model)
                sn = get(c_sn)
                loc_name = get(c_loc)
                mach = get(c_mach)
                status_name = get(c_status)
                proj_name = get(c_proj)
                owner_name_raw = get(c_owner)
                resp_name_raw = get(c_resp)
                notes = get(c_obs)
                image_ref = get(c_img)

                # --- Split inteligente ---
                owner_name = _normalize_name(owner_name_raw)
                resp_name = _normalize_name(resp_name_raw)
                if owner_name and not resp_name:
                    parts = [p for p in _SPLIT_RE.split(owner_name) if p]
                    if len(parts) > 1:
                        owner_name, resp_name = parts[0], parts[1]
                else:
                    if resp_name:
                        parts_r = [p for p in _SPLIT_RE.split(resp_name) if p]
                        if parts_r:
                            resp_name = parts_r[0]

                # Associações com auto-criação
                owner = _get_or_create_user_by_fullname(owner_name, counters) if owner_name else None
                resp = _get_or_create_user_by_fullname(resp_name, counters) if resp_name else None
                loc = _get_or_create_client_by_name(loc_name, counters) if loc_name else None

                proj = None
                if proj_name:
                    proj = Project.query.filter(db.func.lower(Project.name)==proj_name.lower()).first()
                    if not proj:
                        proj = Project(name=proj_name)
                        db.session.add(proj)
                        db.session.flush()

                st = None
                if status_name:
                    st = EquipmentStatus.query.filter(db.func.lower(EquipmentStatus.name)==status_name.lower()).first()
                    if not st:
                        st = EquipmentStatus(name=status_name)
                        db.session.add(st)
                        db.session.flush()

                # Dedup
                existing = None
                if sn:
                    existing = Equipment.query.filter_by(serial_number=sn).first()
                if not existing and pn and name:
                    existing = Equipment.query.filter_by(pn=pn, name=name).first()

                if existing:
                    existing.pn = pn or existing.pn
                    existing.asset_tag = existing.asset_tag or pn or existing.pn
                    existing.model_number = model_number or existing.model_number
                    existing.machine_installed = mach or existing.machine_installed
                    existing.image_ref = image_ref or existing.image_ref
                    existing.location_id = (loc.id if loc else existing.location_id)
                    existing.project_id = (proj.id if proj else existing.project_id)
                    existing.status_id = (st.id if st else existing.status_id)
                    existing.owner_id = (owner.id if owner else existing.owner_id)
                    existing.current_responsible_id = (resp.id if resp else existing.current_responsible_id)
                    existing.notes = '\n'.join([v for v in [existing.notes or '', notes] if v]).strip() or None

                    updated += 1
                else:
                    e = Equipment(
                        name=name,
                        pn=pn or None,
                        asset_tag=pn or None,
                        model_number=model_number or None,
                        serial_number=sn or None,
                        machine_installed=mach or None,
                        image_ref=image_ref or None,
                        location_id=(loc.id if loc else None),
                        project_id=(proj.id if proj else None),
                        status_id=(st.id if st else None),
                        owner_id=(owner.id if owner else None),
                        current_responsible_id=(resp.id if resp else None),
                        notes=notes or None,
                    )
                    db.session.add(e)
                    created += 1

            db.session.commit()

            msg = (f"Importação: {created} criado(s), {updated} atualizado(s). "
                   f"Auto-criados: {counters.get('users',0)} usuário(s), {counters.get('clients',0)} cliente(s).")
            if missing_users:
                msg += f" (Nomes sem correspondência exata: {', '.join(sorted(missing_users))})"
            flash(msg, 'success')
            return redirect(url_for('inventory.list_'))
        except Exception as e:
            flash(f'Falha ao importar: {e}', 'error')
            return redirect(request.url)
    return render_template('inventory/import.html')


def _fill_choices(form: EquipmentForm):
    users = User.query.order_by(User.full_name.asc()).all()
    form.owner_id.choices = [(u.id, u.full_name) for u in users]
    form.owner_id.choices.insert(0, (0, "— Selecione —"))
    form.current_responsible_id.choices = [(u.id, u.full_name) for u in users]
    form.current_responsible_id.choices.insert(0, (0, "— Selecione —"))

    clients = Client.query.order_by(Client.nome_razao.asc()).all()
    form.location_id.choices = [(c.id, c.nome_razao) for c in clients]
    form.location_id.choices.insert(0, (0, "— Selecione —"))

    projects = Project.query.order_by(Project.name.asc()).all()
    form.project_id.choices = [(p.id, p.name) for p in projects]
    form.project_id.choices.insert(0, (0, "— Selecione —"))

    statuses = EquipmentStatus.query.order_by(EquipmentStatus.name.asc()).all()
    form.status_id.choices = [(s.id, s.name) for s in statuses]
    form.status_id.choices.insert(0, (0, "— Selecione —"))
